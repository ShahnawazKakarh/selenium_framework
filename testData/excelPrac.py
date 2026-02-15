import openpyxl

book = openpyxl.load_workbook('/Volumes/Home/python/selenium_framework/testData/PythonDemo.xlsx')

sheet = book.active
Dict = {}
cell = sheet.cell(row=1, column=2)
print(cell.value)

sheet.cell(row=2, column=2, value='Hello')
print(sheet.cell(row=2, column=2).value)

print(sheet.max_row)
print(sheet.max_column)

print(sheet['A5'].value)

for row in range(1, sheet.max_row + 1):
    for column in range(1, sheet.max_column + 1):
        print(sheet.cell(row=row, column=column).value)

# Getting specific values of TestCase2 and adding them in Dict
for row in range(1, sheet.max_row + 1):
    if sheet.cell(row=row, column=1).value == 'Testcase2':
        for column in range(2, sheet.max_column + 1):
            Dict[sheet.cell(row=1, column=column).value] = sheet.cell(row=row, column=column).value
